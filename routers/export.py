import csv
import io
import logging
import uuid
from datetime import datetime
from decimal import Decimal
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from database import get_db
# CHANGE: Import get_current_user instead of require_manager
from middleware.auth import get_current_user
from models.user import User
from models.goal import Goal
from models.goal_sheet import GoalSheet
from models.achievement import Achievement
from models.enums import UserRole
from schemas.user import CurrentUser

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["Export"])

EXPORT_COLUMNS = [
    "Employee Name", "Department", "Manager", "Thrust Area",
    "Goal Title", "UoM", "Target", "Actual", "Score",
    "Weightage", "Quarter", "Financial Year",
]


async def _get_export_data(
    db: AsyncSession,
    current_user: CurrentUser,
    financial_year: str,
    quarter: str | None = None,
) -> list[dict]:
    """Fetch achievement data with role-based filtering."""
    query = (
        select(
            User.full_name.label("employee_name"),
            User.department,
            GoalSheet.financial_year,
            Goal.thrust_area,
            Goal.title.label("goal_title"),
            Goal.uom_type,
            Goal.target_value,
            Goal.target_date,
            Goal.weightage,
            Achievement.actual_value,
            Achievement.actual_date,
            Achievement.score,
            Achievement.quarter,
        )
        .join(GoalSheet, GoalSheet.employee_id == User.id)
        .join(Goal, Goal.sheet_id == GoalSheet.id)
        .outerjoin(Achievement, Achievement.goal_id == Goal.id)
        .where(GoalSheet.financial_year == financial_year)
    )

    if quarter:
        query = query.where(Achievement.quarter == quarter)

    # This handles the secure data scoping perfectly!
    if current_user.role == UserRole.manager:
        query = query.where(User.manager_id == current_user.id)
    elif current_user.role == UserRole.employee:
        query = query.where(User.id == current_user.id)

    query = query.order_by(User.department, User.full_name, Goal.order_index)

    result = await db.execute(query)
    rows = []

    for row in result.all():
        manager_name = ""
        if True:  
            manager_name = "—"

        target_display = ""
        if row.uom_type.value in ("numeric", "percentage"):
            target_display = str(row.target_value) if row.target_value else ""
        elif row.uom_type.value == "timeline":
            target_display = str(row.target_date) if row.target_date else ""
        elif row.uom_type.value == "zero_based":
            target_display = "0"

        actual_display = ""
        if row.actual_value is not None:
            actual_display = str(row.actual_value)
        elif row.actual_date is not None:
            actual_display = str(row.actual_date)

        rows.append({
            "Employee Name": row.employee_name,
            "Department": row.department or "—",
            "Manager": manager_name,
            "Thrust Area": row.thrust_area,
            "Goal Title": row.goal_title,
            "UoM": row.uom_type.value,
            "Target": target_display,
            "Actual": actual_display,
            "Score": str(row.score) if row.score else "—",
            "Weightage": str(row.weightage),
            "Quarter": row.quarter.value if row.quarter else "—",
            "Financial Year": row.financial_year,
        })

    return rows


@router.get(
    "/achievements",
    summary="Export achievement data as CSV or XLSX",
)
async def export_achievements(
    # CHANGE: Replaced require_manager with get_current_user so all roles can access
    current_user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    financial_year: str = Query(default="FY2025-26"),
    quarter: str | None = None,
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
):
    """
    Export achievement data as CSV or XLSX file.
    Admin sees all data; Manager sees only their team; Employee sees only their own.
    """
    rows = await _get_export_data(db, current_user, financial_year, quarter)

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for the specified filters.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"achievements_{financial_year}_{timestamp}"

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    elif format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Achievements"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                cell.border = thin_border

        for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
            max_length = len(header)
            for row_idx in range(2, len(rows) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 3, 40)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )